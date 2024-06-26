import os
from xvfbwrapper import Xvfb

# Start the virtual display
vdisplay = Xvfb()
vdisplay.start()

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import load_workbook
import pywhatkit as kit
import pyautogui as pg
import time
import warnings
import streamlit as st
import pandas as pd

warnings.filterwarnings("ignore")

def apply_dark_mode():
    dark_mode_css = """
    <style>
    /* Set dark background and text color */
    .css-1d391kg, .css-12oz5g7, .css-1y4p8pa {
        background-color: #0e1117;
        color: #ffffff;
    }
    /* Sidebar background color */
    .css-1d3fmxh {
        background-color: #0e1117;
    }
    /* Adjust text color */
    .css-17eq0hr {
        color: #ffffff;
    }
    </style>
    """
    st.markdown(dark_mode_css, unsafe_allow_html=True)

apply_dark_mode()

# SMTP configuration
your_name = "Sekolah Harapan Bangsa"
your_email = "shsmodernhill@shb.sch.id"
your_password = "jvvmdgxgdyqflcrf"

server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
server.ehlo()
server.login(your_email, your_password)

ALLOWED_EXTENSIONS = {'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def send_whatsapp_messages(data, announcement=False, invoice=False, proof_payment=False):
    # Open WhatsApp Web
    kit.sendwhatmsg("+0000000000", "", 0, 0, wait_time=20, tab_close=True)  # Dummy message to open WhatsApp Web
    st.info("Please scan the QR code in the opened WhatsApp Web window.")
    time.sleep(45)

    for index, row in data.iterrows():
        phone_number = str(row['Phone Number'])
        if not phone_number.startswith('+62'):
            phone_number = f'+62{phone_number.lstrip("0")}'

        if announcement:
            message = f"""
            Kepada Yth. Orang Tua/Wali Murid *{row['Nama_Siswa']}*,
            Kami hendak menyampaikan info mengenai:
            *Subject:* {row['Subject']}
            *Description:* {row['Description']}
            *Link:* {row['Link']}
            Terima kasih atas kerjasamanya.
            Admin Sekolah
            
            Jika ada pertanyaan atau hendak konfirmasi dapat menghubungi:
            • Ibu Penna (Kasir): https://bit.ly/mspennashb
            • Bapak Supatmin (Admin SMP & SMA): https://bit.ly/wamrsupatminshb4
            """
        elif invoice:
            message = f"""
            Kepada Yth. Orang Tua/Wali Murid *{row['customer_name']}* (Kelas *{row['Grade']}*),
            Kami hendak menyampaikan info mengenai:
            • *Subject:* {row['Subject']}
            • *Batas Tanggal Pembayaran:* {row['expired_date']}
            • *Sebesar:* Rp. {row['trx_amount']:,.2f}
            • Pembayaran via nomor *virtual account* (VA) BNI/Bank: *{row['virtual_account']}*
        Terima kasih atas kerjasamanya.
        Admin Sekolah
        Jika ada pertanyaan atau hendak konfirmasi dapat menghubungi:
            • Ibu Penna (Kasir): https://bit.ly/mspennashb
            • Bapak Supatmin (Admin SMP & SMA): https://bit.ly/wamrsupatminshb4
            """
        elif proof_payment:
            message = f"""
            Kepada Yth. Orang Tua/Wali Murid *{row['Nama_Siswa']}* (Kelas *{row['Grade']}*),
            Kami hendak menyampaikan info mengenai SPP:
            • *SPP yang sedang berjalan:* {row['bulan_berjalan']:,.2f} ({row['Ket_1']})
            • *Denda:* {row['Denda']:,.2f} ({row['Ket_3']})
            • *SPP bulan-bulan sebelumnya:* {row['SPP_30hari']:,.2f} ({row['Ket_2']})
            • *Keterangan:* {row['Ket_4']}
            • *Total tagihan:* {row['Total']:,.2f}
            Terima kasih atas kerjasamanya.
            Admin Sekolah
            
            Jika ada pertanyaan atau hendak konfirmasi dapat menghubungi:
            • Ibu Penna (Kasir): https://bit.ly/mspennashb
            • Bapak Supatmin (Admin SMP & SMA): https://bit.ly/wamrsupatminshb4
            """
        else:
            continue

        while True:
            try:
                kit.sendwhatmsg_instantly(phone_number, message, wait_time=20)
                time.sleep(20)
                st.success(f"Message sent successfully to {phone_number}")
                break
            except Exception as e:
                st.error(f"Failed to send message to {phone_number}: {str(e)}. Retrying...")
                time.sleep(20)

def send_emails(email_list, announcement=False, invoice=False, proof_payment=False):
    for idx, entry in enumerate(email_list):
        if announcement:
            subject = entry['Subject']
            name = entry['Nama_Siswa']
            email = entry['Email']
            description = entry['Description']
            link = entry['Link']
            message = f"""
            Kepada Yth.<br>Orang Tua/Wali Murid <span style="color: #007bff;">{name}</span><br>
            <p>Salam Hormat,</p>
            <p>Kami hendak menyampaikan info mengenai:</p>
            <ul>
                <li><strong>Subject:</strong> {subject}</li>
                <li><strong>Description:</strong> {description}</li>
                <li><strong>Link:</strong> {link}</li>
            </ul>
            <p>Terima kasih atas kerjasamanya.</p>
            <p>Admin Sekolah</p>
            <p>Jika ada pertanyaan atau hendak konfirmasi dapat menghubungi:</p>
            <strong>Ibu Penna (Kasir):</strong> <a href='https://bit.ly/mspennashb' style="color: #007bff;">https://bit.ly/mspennashb</a><br>
            <strong>Bapak Supatmin (Admin SMP & SMA):</strong> <a href='https://bit.ly/wamrsupatminshb4' style="color: #007bff;">https://bit.ly/wamrsupatminshb4</a>
            """
        elif invoice:
            subject = entry['Subject']
            grade = entry['Grade']
            va = entry['virtual_account']
            name = entry['customer_name']
            email = entry['customer_email']
            nominal = "{:,.2f}".format(entry['trx_amount'])
            expired_date = entry['expired_date']
            expired_time = entry['expired_time']
            description = entry['description']
            link = entry['link']
            message = f"""
            Kepada Yth.<br>Orang Tua/Wali Murid <span style="color: #007bff;">{name}</span> (Kelas <span style="color: #007bff;">{grade}</span>)<br>
            <p>Salam Hormat,</p>
            <p>Kami hendak menyampaikan info mengenai:</p>
            <ul>
                <li><strong>Subject:</strong> {subject}</li>
                <li><strong>Batas Tanggal Pembayaran:</strong> {expired_date}</li>
                <li><strong>Sebesar:</strong> Rp. {nominal}</li>
                <li><strong>Pembayaran via nomor virtual account (VA) BNI/Bank:</strong> {va}</li>
            </ul>
            <p>Terima kasih atas kerjasamanya.</p>
            <p>Admin Sekolah</p>
            <p>Jika ada pertanyaan atau hendak konfirmasi dapat menghubungi:</p>
            <strong>Ibu Penna (Kasir):</strong> <a href='https://bit.ly/mspennashb' style="color: #007bff;">https://bit.ly/mspennashb</a><br>
            <strong>Bapak Supatmin (Admin SMP & SMA):</strong> <a href='https://bit.ly/wamrsupatminshb4' style="color: #007bff;">https://bit.ly/wamrsupatminshb4</a>
            """
        elif proof_payment:
            subject = "Bukti Pembayaran"
            name = entry['Nama_Siswa']
            grade = entry['Grade']
            email = entry['Email']
            bulan_berjalan = "{:,.2f}".format(entry['bulan_berjalan'])
            ket_1 = entry['Ket_1']
            spp_30hari = "{:,.2f}".format(entry['SPP_30hari'])
            ket_2 = entry['Ket_2']
            denda = "{:,.2f}".format(entry['Denda'])
            ket_3 = entry['Ket_3']
            ket_4 = entry['Ket_4']
            total = "{:,.2f}".format(entry['Total'])
            message = f"""
            Kepada Yth.<br>Orang Tua/Wali Murid <span style="color: #007bff;">{name}</span> (Kelas <span style="color: #007bff;">{grade}</span>)<br>
            <p>Salam Hormat,</p>
            <p>Kami hendak menyampaikan info mengenai SPP:</p>
            <ul>
                <li><strong>SPP yang sedang berjalan:</strong> {bulan_berjalan} ({ket_1})</li>
                <li><strong>SPP bulan-bulan sebelumnya:</strong> {spp_30hari} ({ket_2})</li>
                <li><strong>Denda:</strong> {denda} ({ket_3})</li>
                <li><strong>Keterangan:</strong> {ket_4}</li>
                <li><strong>Total tagihan:</strong> {total}</li>
            </ul>
            <p>Terima kasih atas kerjasamanya.</p>
            <p>Admin Sekolah</p>
            <p>Jika ada pertanyaan atau hendak konfirmasi dapat menghubungi:</p>
            <strong>Ibu Penna (Kasir):</strong> <a href='https://bit.ly/mspennashb' style="color: #007bff;">https://bit.ly/mspennashb</a><br>
            <strong>Bapak Supatmin (Admin SMP & SMA):</strong> <a href='https://bit.ly/wamrsupatminshb4' style="color: #007bff;">https://bit.ly/wamrsupatminshb4</a>
            """
        else:
            continue

        msg = MIMEMultipart()
        msg['From'] = your_email
        msg['To'] = email
        msg['Subject'] = subject

        msg.attach(MIMEText(message, 'html'))

        try:
            server.sendmail(your_email, email, msg.as_string())
            st.success(f"Email sent successfully to {name}")
        except Exception as e:
            st.error(f"Failed to send email to {name}: {str(e)}")

def process_uploaded_file(uploaded_file):
    if uploaded_file is not None and allowed_file(uploaded_file.name):
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active

        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {ws.cell(row=1, column=col_idx).value: cell_value for col_idx, cell_value in enumerate(row, start=1)}
            data.append(row_data)

        return pd.DataFrame(data)
    return None

st.title('Sekolah Harapan Bangsa')
st.header('Welcome to the email and WhatsApp blast tool')

uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])

if uploaded_file is not None:
    st.write("File uploaded successfully!")

    data = process_uploaded_file(uploaded_file)

    if data is not None:
        st.write(data)

        email_announcement = st.checkbox("Email Announcement")
        email_invoice = st.checkbox("Email Invoice")
        email_proof_payment = st.checkbox("Email Proof of Payment")

        if st.button("Send Emails"):
            if email_announcement:
                send_emails(data.to_dict('records'), announcement=True)
            elif email_invoice:
                send_emails(data.to_dict('records'), invoice=True)
            elif email_proof_payment:
                send_emails(data.to_dict('records'), proof_payment=True)
            else:
                st.warning("Please select an email option.")

        whatsapp_announcement = st.checkbox("WhatsApp Announcement")
        whatsapp_invoice = st.checkbox("WhatsApp Invoice")
        whatsapp_proof_payment = st.checkbox("WhatsApp Proof of Payment")

        if st.button("Send WhatsApp Messages"):
            if whatsapp_announcement:
                send_whatsapp_messages(data, announcement=True)
            elif whatsapp_invoice:
                send_whatsapp_messages(data, invoice=True)
            elif whatsapp_proof_payment:
                send_whatsapp_messages(data, proof_payment=True)
            else:
                st.warning("Please select a WhatsApp message option.")
    else:
        st.error("Failed to process the uploaded file. Please ensure it is a valid Excel file.")
else:
    st.info("Please upload an Excel file.")

# Stop the virtual display
vdisplay.stop()

# Close the SMTP server connection
server.quit()
